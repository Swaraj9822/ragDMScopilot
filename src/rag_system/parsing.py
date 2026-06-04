"""Document parsing — multi-format support for the RAG ingestion pipeline.

Supported formats:
    LlamaParse  : PDF, DOCX, DOC, PPTX, PPT, ODT, RTF, EPUB, XML, XLS,
                  images (JPG, PNG, TIFF, BMP, WebP, GIF)
    Spreadsheet : XLSX, CSV
    HTML        : HTML, HTM
    Text        : TXT, MD, Markdown, RST
"""

from __future__ import annotations

import csv
import io
from pathlib import Path
from tempfile import NamedTemporaryFile

from llama_cloud_services import LlamaParse

from rag_system.config import Settings
from rag_system.models import ParsedDocument
from rag_system.observability import get_logger, metrics, retry_on_transient

logger = get_logger(__name__)

# ---------------------------------------------------------------------------
# Supported format groups
# ---------------------------------------------------------------------------

LLAMAPARSE_EXTENSIONS = frozenset({
    # Office documents
    ".pdf", ".docx", ".doc", ".pptx", ".ppt", ".odt", ".rtf", ".epub",
    # Spreadsheets handled by LlamaParse (XLS legacy; XLSX/CSV go to SpreadsheetParser)
    ".xls",
    # Structured data
    ".xml",
    # Images (OCR via LlamaParse)
    ".jpg", ".jpeg", ".png", ".tiff", ".tif", ".bmp", ".webp", ".gif",
})
SPREADSHEET_EXTENSIONS = frozenset({".xlsx", ".csv"})
HTML_EXTENSIONS = frozenset({".html", ".htm"})
TEXT_EXTENSIONS = frozenset({".txt", ".md", ".markdown", ".rst"})

SUPPORTED_EXTENSIONS = LLAMAPARSE_EXTENSIONS | SPREADSHEET_EXTENSIONS | HTML_EXTENSIONS | TEXT_EXTENSIONS


def detect_parser_type(filename: str) -> str:
    """Return the parser category for a given filename, or raise ValueError."""
    ext = Path(filename).suffix.lower()
    if ext in LLAMAPARSE_EXTENSIONS:
        return "llamaparse"
    if ext in SPREADSHEET_EXTENSIONS:
        return "spreadsheet"
    if ext in HTML_EXTENSIONS:
        return "html"
    if ext in TEXT_EXTENSIONS:
        return "text"
    raise ValueError(
        f"Unsupported file format: '{ext}'. "
        f"Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
    )


# ---------------------------------------------------------------------------
# LlamaParse — PDF, DOCX, PPTX, images (OCR)
# ---------------------------------------------------------------------------


class LlamaParseDocumentParser:
    """Parses complex documents using LlamaParse (PDF, DOCX, PPTX, images)."""

    def __init__(self, settings: Settings):
        self._parser = LlamaParse(
            api_key=settings.llama_cloud_api_key,
            result_type="markdown",
            premium_mode=True,
        )

    @retry_on_transient()
    async def parse(
        self, document_id: str, version: str, filename: str, content: bytes
    ) -> ParsedDocument:
        suffix = Path(filename).suffix or ".pdf"
        with NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        logger.info(
            "Sending %s to LlamaParse (%d bytes)",
            filename,
            len(content),
            extra={"document_id": document_id, "version": version, "file_name": filename},
        )
        try:
            documents = await self._parser.aload_data(tmp_path)
        finally:
            Path(tmp_path).unlink(missing_ok=True)

        markdown = "\n\n".join(doc.text for doc in documents)
        logger.info(
            "LlamaParse returned %d page(s), %d chars of markdown",
            len(documents),
            len(markdown),
            extra={"document_id": document_id, "version": version},
        )
        return ParsedDocument(
            document_id=document_id,
            version=version,
            markdown=markdown,
            metadata={"source_filename": filename, "parser": "llamaparse"},
        )


# Backward-compatible alias
LlamaParsePdfParser = LlamaParseDocumentParser


# ---------------------------------------------------------------------------
# Spreadsheet parser — XLSX, CSV
# ---------------------------------------------------------------------------


class SpreadsheetParser:
    """Converts XLSX and CSV files into markdown tables."""

    async def parse(
        self, document_id: str, version: str, filename: str, content: bytes
    ) -> ParsedDocument:
        ext = Path(filename).suffix.lower()
        logger.info(
            "Parsing spreadsheet %s (%d bytes)",
            filename,
            len(content),
            extra={"document_id": document_id, "version": version, "file_name": filename},
        )

        if ext == ".csv":
            markdown = self._parse_csv(content)
        else:
            markdown = self._parse_xlsx(content)

        logger.info(
            "Spreadsheet parsed: %d chars of markdown",
            len(markdown),
            extra={"document_id": document_id, "version": version},
        )
        return ParsedDocument(
            document_id=document_id,
            version=version,
            markdown=markdown,
            metadata={"source_filename": filename, "parser": "spreadsheet"},
        )

    def _parse_xlsx(self, content: bytes) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "Install openpyxl to parse XLSX files: pip install openpyxl"
            ) from exc

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sections: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            sections.append(f"## {sheet_name}\n")
            sections.append(_rows_to_markdown_table(rows))

        wb.close()
        return "\n\n".join(sections) if sections else "*(Empty spreadsheet)*"

    def _parse_csv(self, content: bytes) -> str:
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = [tuple(row) for row in reader]
        if not rows:
            return "*(Empty CSV)*"
        return _rows_to_markdown_table(rows)


def _rows_to_markdown_table(rows: list[tuple]) -> str:
    """Convert a list of row tuples into a markdown table string."""
    if not rows:
        return ""

    header = rows[0]
    col_count = len(header)
    header_cells = [str(cell) if cell is not None else "" for cell in header]
    lines = [
        "| " + " | ".join(header_cells) + " |",
        "| " + " | ".join("---" for _ in range(col_count)) + " |",
    ]
    for row in rows[1:]:
        cells = [str(cell) if cell is not None else "" for cell in row]
        # Pad or truncate to match header column count
        cells = (cells + [""] * col_count)[:col_count]
        lines.append("| " + " | ".join(cells) + " |")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# HTML parser
# ---------------------------------------------------------------------------


class HtmlParser:
    """Extracts text content from HTML and converts to structured markdown."""

    _HEADING_MAP = {
        "h1": "#", "h2": "##", "h3": "###",
        "h4": "####", "h5": "#####", "h6": "######",
    }

    async def parse(
        self, document_id: str, version: str, filename: str, content: bytes
    ) -> ParsedDocument:
        logger.info(
            "Parsing HTML %s (%d bytes)",
            filename,
            len(content),
            extra={"document_id": document_id, "version": version, "file_name": filename},
        )

        try:
            from bs4 import BeautifulSoup
        except ImportError as exc:
            raise RuntimeError(
                "Install beautifulsoup4 to parse HTML files: pip install beautifulsoup4"
            ) from exc

        text = content.decode("utf-8", errors="replace")
        soup = BeautifulSoup(text, "html.parser")

        # Remove non-content elements
        for tag in soup(["script", "style", "nav", "footer", "header", "noscript"]):
            tag.decompose()

        title = soup.title.string.strip() if soup.title and soup.title.string else None
        markdown = self._soup_to_markdown(soup)

        logger.info(
            "HTML parsed: %d chars of markdown",
            len(markdown),
            extra={"document_id": document_id, "version": version},
        )
        metadata: dict = {"source_filename": filename, "parser": "html"}
        if title:
            metadata["title"] = title

        return ParsedDocument(
            document_id=document_id,
            version=version,
            markdown=markdown,
            metadata=metadata,
        )

    def _soup_to_markdown(self, soup) -> str:
        """Walk the DOM and produce structured markdown."""
        lines: list[str] = []
        content_tags = [
            "h1", "h2", "h3", "h4", "h5", "h6",
            "p", "li", "pre", "blockquote",
        ]

        for element in soup.find_all(content_tags):
            text = element.get_text(separator=" ", strip=True)
            if not text:
                continue
            tag = element.name
            if tag in self._HEADING_MAP:
                lines.append(f"\n{self._HEADING_MAP[tag]} {text}\n")
            elif tag == "li":
                lines.append(f"- {text}")
            elif tag == "pre":
                lines.append(f"\n```\n{text}\n```\n")
            elif tag == "blockquote":
                lines.append(f"> {text}")
            else:
                lines.append(f"\n{text}\n")

        result = "\n".join(lines).strip()
        if not result:
            # Fallback: extract all text if structured parsing yields nothing
            result = soup.get_text(separator="\n", strip=True)
        return result


# ---------------------------------------------------------------------------
# Plain text / Markdown parser
# ---------------------------------------------------------------------------


class TextParser:
    """Handles plain text and markdown files directly."""

    async def parse(
        self, document_id: str, version: str, filename: str, content: bytes
    ) -> ParsedDocument:
        logger.info(
            "Parsing text file %s (%d bytes)",
            filename,
            len(content),
            extra={"document_id": document_id, "version": version, "file_name": filename},
        )

        text = content.decode("utf-8", errors="replace")
        ext = Path(filename).suffix.lower()
        parser_name = "markdown" if ext in (".md", ".markdown") else "text"

        logger.info(
            "Text file parsed: %d chars",
            len(text),
            extra={"document_id": document_id, "version": version},
        )
        return ParsedDocument(
            document_id=document_id,
            version=version,
            markdown=text,
            metadata={"source_filename": filename, "parser": parser_name},
        )


# ---------------------------------------------------------------------------
# Document parser router
# ---------------------------------------------------------------------------


class DocumentParserRouter:
    """Routes documents to the appropriate parser based on file extension.

    Dispatching:
        PDF / DOCX / PPTX / images  →  LlamaParse (cloud)
        XLSX / CSV                  →  openpyxl / csv  (local)
        HTML / HTM                  →  BeautifulSoup   (local)
        TXT / MD / RST              →  direct read     (local)
    """

    def __init__(self, settings: Settings):
        self._llamaparse = LlamaParseDocumentParser(settings)
        self._spreadsheet = SpreadsheetParser()
        self._html = HtmlParser()
        self._text = TextParser()
        logger.info(
            "DocumentParserRouter initialised (supported_formats=%d)",
            len(SUPPORTED_EXTENSIONS),
        )

    async def parse(
        self, document_id: str, version: str, filename: str, content: bytes
    ) -> ParsedDocument:
        """Parse a document by routing to the correct parser based on file extension."""
        parser_type = detect_parser_type(filename)
        logger.info(
            "Routing %s to '%s' parser",
            filename,
            parser_type,
            extra={
                "document_id": document_id,
                "version": version,
                "file_name": filename,
                "parser_type": parser_type,
            },
        )
        metrics.increment("rag_documents_parsed_total", {"parser": parser_type})

        if parser_type == "llamaparse":
            return await self._llamaparse.parse(document_id, version, filename, content)
        elif parser_type == "spreadsheet":
            return await self._spreadsheet.parse(document_id, version, filename, content)
        elif parser_type == "html":
            return await self._html.parse(document_id, version, filename, content)
        else:
            return await self._text.parse(document_id, version, filename, content)
